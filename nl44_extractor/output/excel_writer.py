"""
Excel Writer for NL-44 (Motor TP Obligations).

Master_Data: two rows per company per quarter — one for "For the Quarter",
one for "Upto the Quarter" — with the five NL-44 item values as columns.
This mirrors the Quarter_Info pattern used in NL-39.

Verification sheet: single table matching the PDF layout —
  Rows = items, Cols = For the Quarter | Upto the Quarter.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config.settings import (
    MASTER_COLUMNS, EXTRACTOR_VERSION, NUMBER_FORMAT,
    LOW_CONFIDENCE_FILL_COLOR, company_key_to_pascal,
)
from config.row_registry import ROW_ORDER, ROW_DISPLAY_NAMES
from config.company_metadata import get_metadata
from extractor.models import CompanyExtract

logger = logging.getLogger(__name__)

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_META_FILL    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_QTR_FILL     = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_YTD_FILL     = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

_METRIC_COLUMNS = {c for c in MASTER_COLUMNS if c.lower() in ROW_ORDER}


def _year_code_to_fy_end(year_code: str) -> str:
    s = str(year_code).strip()
    if len(s) == 8:
        return s[4:]
    if len(s) == 6:
        return f"20{s[4:]}"
    return s


# ---------------------------------------------------------------------------
# Master_Data sheet
# ---------------------------------------------------------------------------

def _write_master_data(ws, extractions: List[CompanyExtract],
                       existing_rows: Optional[List[list]] = None):
    for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
    ws.freeze_panes = "A2"

    current_row = 2

    if existing_rows:
        for row_data in existing_rows:
            for col_idx, val in enumerate(row_data, 1):
                if col_idx > len(MASTER_COLUMNS):
                    break
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                if MASTER_COLUMNS[col_idx - 1] in _METRIC_COLUMNS:
                    cell.number_format = NUMBER_FORMAT
            current_row += 1

    for extract in extractions:
        meta = get_metadata(extract.company_key)
        period_data = extract.current_year
        if not period_data:
            continue

        for quarter_info, p_key in [("For the Quarter", "qtr"), ("Upto the Quarter", "ytd")]:
            metadata = {
                "Company_Name":         meta["company_name"],
                "Company":              meta["sorted_company"],
                "NL":                   extract.form_type,
                "Quarter":              extract.quarter,
                "Year":                 _year_code_to_fy_end(extract.year),
                "Quarter_Info":         quarter_info,
                "Sector":               meta["sector"],
                "Industry_Competitors": meta["competitors"],
                "GI_Companies":         "GI Company",
                "Source_File":          extract.source_file,
            }

            row_values = []
            for col_name in MASTER_COLUMNS:
                if col_name in metadata:
                    row_values.append(metadata[col_name])
                elif col_name.lower() in ROW_ORDER:
                    val = period_data.data.get(col_name.lower(), {}).get(p_key)
                    row_values.append(val)
                else:
                    row_values.append(None)

            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                if MASTER_COLUMNS[col_idx - 1] in _METRIC_COLUMNS:
                    cell.number_format = NUMBER_FORMAT
            current_row += 1


# ---------------------------------------------------------------------------
# Verification sheet — matches PDF layout
# Rows = items, Cols = For the Quarter | Upto the Quarter
# ---------------------------------------------------------------------------

def _write_verification_sheet(ws, extract: CompanyExtract):
    ws.cell(row=1, column=1, value=f"VERIFICATION SHEET: {extract.company_name}") \
      .font = Font(bold=True, size=14)
    ws.cell(row=2, column=1,
            value=f"Quarter: {extract.quarter} | Year: {extract.year} | Source: {extract.source_file}")

    if not extract.current_year:
        ws.cell(row=4, column=1, value="No data extracted.").font = Font(italic=True)
        return

    _write_pdf_table(ws, extract.current_year, start_row=4)


def _write_pdf_table(ws, period_data, start_row: int):
    """
    PDF-matching layout:
      Row start_row:    title
      Row start_row+1:  header — Items | For the Quarter | Upto the Quarter
      Rows start_row+2+: one row per item
    """
    # Title
    title_cell = ws.cell(row=start_row, column=1,
                         value="Motor TP Obligations (Quarterly Returns)")
    title_cell.font = Font(bold=True, color="FFFFFF", size=11)
    title_cell.fill = _HEADER_FILL
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=3)
    title_cell.alignment = _CENTER_ALIGN

    # Header row
    h = start_row + 1
    for col_idx, label in enumerate(["Items", "For the Quarter", "Upto the Quarter"], 1):
        cell = ws.cell(row=h, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    # Data rows
    for r_idx, metric_key in enumerate(ROW_ORDER):
        ws_row = h + 1 + r_idx
        label_cell = ws.cell(row=ws_row, column=1,
                             value=ROW_DISPLAY_NAMES.get(metric_key, metric_key))

        qtr_val = period_data.data.get(metric_key, {}).get("qtr")
        ytd_val = period_data.data.get(metric_key, {}).get("ytd")

        qtr_cell = ws.cell(row=ws_row, column=2, value=qtr_val)
        ytd_cell = ws.cell(row=ws_row, column=3, value=ytd_val)
        qtr_cell.number_format = NUMBER_FORMAT
        ytd_cell.number_format = NUMBER_FORMAT
        qtr_cell.fill = _QTR_FILL
        ytd_cell.fill = _YTD_FILL


# ---------------------------------------------------------------------------
# Meta sheet
# ---------------------------------------------------------------------------

def _write_meta_sheet(ws, extractions: List[CompanyExtract], stats: Dict[str, Any]):
    companies = sorted({e.company_name for e in extractions})
    quarters  = sorted({f"{e.quarter}_{e.year}" for e in extractions})

    data = [
        ["Key", "Value"],
        ["extraction_date",   datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["extractor_version", EXTRACTOR_VERSION],
        ["files_processed",   stats.get("files_processed", 0)],
        ["files_succeeded",   stats.get("files_succeeded", 0)],
        ["files_failed",      stats.get("files_failed", 0)],
        ["companies",         ", ".join(companies)],
        ["quarters",          ", ".join(quarters)],
    ]
    for r_idx, row in enumerate(data, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
            else:
                cell.fill = _META_FILL


def _sheet_name_for(extract: CompanyExtract) -> str:
    name = f"{company_key_to_pascal(extract.company_key)}_{extract.quarter}_{extract.year}"
    return name[:31]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def save_workbook(extractions: List[CompanyExtract], output_path: str,
                  stats: Optional[Dict[str, Any]] = None):
    if stats is None:
        stats = {}

    output_file = Path(output_path)
    existing_rows = []

    if output_file.exists():
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(output_path)
        new_files = {e.source_file for e in extractions}

        if "Master_Data" in wb.sheetnames:
            ws_old = wb["Master_Data"]
            headers = [cell.value for cell in ws_old[1]]
            if headers[:len(MASTER_COLUMNS)] == MASTER_COLUMNS:
                try:
                    sf_idx = headers.index("Source_File")
                except ValueError:
                    sf_idx = None
                if sf_idx is not None:
                    for row in ws_old.iter_rows(min_row=2, values_only=True):
                        if row[sf_idx] and row[sf_idx] not in new_files:
                            existing_rows.append(list(row))
            else:
                logger.warning("Existing Master_Data has different column layout — regenerating.")
            del wb["Master_Data"]

        for extract in extractions:
            sn = _sheet_name_for(extract)
            if sn in wb.sheetnames:
                del wb[sn]
        if "_meta" in wb.sheetnames:
            del wb["_meta"]
    else:
        wb = Workbook()
        wb.remove(wb.active)

    ws_master = wb.create_sheet("Master_Data", 0)
    _write_master_data(ws_master, extractions, existing_rows=existing_rows)

    for extract in extractions:
        ws = wb.create_sheet(title=_sheet_name_for(extract))
        _write_verification_sheet(ws, extract)

    ws_meta = wb.create_sheet(title="_meta")
    _write_meta_sheet(ws_meta, extractions, stats)

    wb.save(output_path)
    logger.info(f"Excel workbook saved to {output_path}")


def write_validation_summary_sheet(report_path: str, master_path: str):
    import pandas as pd
    df = pd.read_csv(report_path)
    summary = df.pivot_table(
        index=["company", "quarter", "year"],
        columns="status", aggfunc="size", fill_value=0,
    ).reset_index()
    for col in ["PASS", "WARN", "FAIL", "SKIP"]:
        if col not in summary.columns:
            summary[col] = 0
    summary["Total_Checks"] = summary[["PASS", "SKIP", "WARN", "FAIL"]].sum(axis=1)
    summary = summary.rename(columns={"company": "Company", "quarter": "Quarter", "year": "Year"})
    cols = ["Company", "Quarter", "Year", "Total_Checks", "PASS", "SKIP", "WARN", "FAIL"]
    summary = summary[cols]
    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as w:
        summary.to_excel(w, sheet_name="Validation_Summary", index=False)


def write_validation_detail_sheet(report_path: str, master_path: str):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df = pd.read_csv(report_path)
    cols_map = {
        "company": "Company", "quarter": "Quarter", "year": "Year",
        "period": "Period", "check_name": "Check_Name",
        "status": "Status", "expected": "Expected", "actual": "Actual",
        "delta": "Delta", "note": "Note",
    }
    detail = df[df["status"].isin(["FAIL", "WARN"])].copy()
    if detail.empty:
        detail = pd.DataFrame(columns=list(cols_map.values()))
    else:
        detail = detail.rename(columns=cols_map)[list(cols_map.values())]
        detail = detail.sort_values("Status").reset_index(drop=True)

    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as w:
        detail.to_excel(w, sheet_name="Validation_Detail", index=False)

    wb = load_workbook(master_path)
    ws = wb["Validation_Detail"]
    red_fill    = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    status_col = list(cols_map.values()).index("Status") + 1
    for row_idx in range(2, ws.max_row + 1):
        val  = ws.cell(row=row_idx, column=status_col).value
        fill = red_fill if val == "FAIL" else yellow_fill
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    wb.save(master_path)
    logger.info(f"Validation_Detail sheet written to {master_path}")
